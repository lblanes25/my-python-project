import os
import datetime
import pandas as pd
from typing import Dict, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from config_manager import ConfigManager
from data_processor import DataProcessor
from logging_config import setup_logging

logger = setup_logging()


class ConsolidatedReportGenerator:
    """Generates consolidated Excel reports from multiple QA analytics"""

    def __init__(self, output_dir: str = "output"):
        """
        Initialize consolidated report generator

        Args:
            output_dir: Directory for output files
        """
        self.output_dir = output_dir
        self.config_manager = ConfigManager()

        # Create output directory if it doesn't exist
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

    def run_analytics(self, analytic_ids: List[str], source_files: Dict[str, str]) -> Dict[str, Dict]:
        """
        Run multiple analytics and collect results

        Args:
            analytic_ids: List of analytic IDs to run
            source_files: Dictionary mapping analytic IDs to source file paths

        Returns:
            Dictionary of results by analytic ID
        """
        results_by_analytic = {}

        for analytic_id in analytic_ids:
            if analytic_id not in source_files:
                logger.error(f"No source file provided for analytic ID {analytic_id}")
                continue

            try:
                # Get configuration
                config = self.config_manager.get_config(analytic_id)
                logger.info(f"Running analytic {analytic_id}: {config['analytic_name']}")

                # Initialize processor
                processor = DataProcessor(config)

                # Process data
                success, message = processor.process_data(source_files[analytic_id])

                if success:
                    logger.info(f"Analytic {analytic_id} processed successfully")
                    results_by_analytic[analytic_id] = {
                        'config': config,
                        'results': processor.results,
                        'source_file': source_files[analytic_id]
                    }
                else:
                    logger.error(f"Failed to process analytic {analytic_id}: {message}")

            except Exception as e:
                logger.error(f"Error running analytic {analytic_id}: {e}")

        return results_by_analytic

    def generate_consolidated_reports(self, results_by_analytic: Dict[str, Dict]) -> Dict[str, str]:
        """
        Generate consolidated reports for each audit leader across multiple analytics

        Args:
            results_by_analytic: Dictionary of results by analytic ID

        Returns:
            Dictionary mapping audit leaders to their report paths
        """
        # Collect all audit leaders across all analytics
        audit_leaders = set()
        leader_field_by_analytic = {}

        for analytic_id, data in results_by_analytic.items():
            config = data['config']
            results = data['results']

            # Get the field name used for grouping (usually Audit Leader)
            group_by_field = config['reporting']['group_by']
            leader_field_by_analytic[analytic_id] = group_by_field

            # Extract all unique leader values
            if 'detail' in results and group_by_field in results['detail'].columns:
                leaders = results['detail'][group_by_field].unique()
                audit_leaders.update([l for l in leaders if pd.notna(l)])

        # Generate consolidated main report for QA department
        main_report_path = self.generate_consolidated_main_report(results_by_analytic, leader_field_by_analytic,
                                                                  list(audit_leaders))

        # Generate a consolidated report for each audit leader
        reports_by_leader = {}
        for leader in audit_leaders:
            report_path = self._generate_leader_report(leader, results_by_analytic, leader_field_by_analytic)
            if report_path:
                reports_by_leader[leader] = report_path

        # Add main report to the dictionary with a special key
        if main_report_path:
            reports_by_leader["__MAIN_REPORT__"] = main_report_path

        return reports_by_leader

    def generate_consolidated_main_report(self, results_by_analytic: Dict[str, Dict],
                                          leader_field_by_analytic: Dict[str, str],
                                          all_leaders: List[str]) -> str:
        """
        Generate a consolidated main report for QA department-level evaluation

        Args:
            results_by_analytic: Dictionary of results by analytic ID
            leader_field_by_analytic: Dictionary mapping analytic IDs to their leader field names
            all_leaders: List of all audit leaders

        Returns:
            Path to the generated main report
        """
        # Create filename for main report
        timestamp = datetime.datetime.now().strftime("%Y%m%d")
        filename = f"QA_Department_Consolidated_Report_{timestamp}.xlsx"
        output_path = os.path.join(self.output_dir, filename)

        try:
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                try:
                    # Create department-level executive summary
                    self._create_department_summary(writer, results_by_analytic, leader_field_by_analytic, all_leaders)
                except Exception as e:
                    logger.error(f"Error creating department summary: {e}")
                    # Create a basic summary instead
                    self._create_basic_summary(writer, results_by_analytic)

                # Add summaries by analytic
                for analytic_id, data in results_by_analytic.items():
                    try:
                        config = data['config']
                        results = data['results']

                        # Summary sheet for this analytic
                        if 'summary' in results and not results['summary'].empty:
                            sheet_name = f"QA-{analytic_id} Summary"
                            results['summary'].to_excel(writer, sheet_name=sheet_name, index=False)
                            self._adjust_column_widths(writer.sheets[sheet_name])
                    except Exception as e:
                        logger.error(f"Error adding summary for analytic {analytic_id}: {e}")

                try:
                    # Add configuration info for all analytics
                    config_data = []
                    config_data.append({'Parameter': 'DEPARTMENT LEVEL QA ANALYTICS REPORT', 'Value': ''})
                    config_data.append(
                        {'Parameter': 'Date Generated', 'Value': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')})
                    config_data.append({'Parameter': 'Number of Analytics', 'Value': len(results_by_analytic)})
                    config_data.append({'Parameter': 'Number of Audit Leaders', 'Value': len(all_leaders)})
                    config_data.append({'Parameter': '', 'Value': ''})

                    # Add sections for each analytic
                    for analytic_id, data in results_by_analytic.items():
                        config = data['config']
                        config_data.append(
                            {'Parameter': f'--- QA-{analytic_id}: {config["analytic_name"]} ---', 'Value': ''})
                        config_data.append({'Parameter': 'Description',
                                            'Value': config.get('analytic_description', 'No description available')})
                        config_data.append(
                            {'Parameter': 'Error Threshold (%)', 'Value': config['thresholds']['error_percentage']})

                        # Add validation rule summaries
                        for i, validation in enumerate(config.get('validations', []), 1):
                            rule_name = validation.get('rule', 'Unknown')
                            description = validation.get('description', 'No description')
                            config_data.append({'Parameter': f'Rule {i}', 'Value': f"{rule_name}: {description}"})

                        config_data.append({'Parameter': '', 'Value': ''})

                    # Write configuration data
                    pd.DataFrame(config_data).to_excel(writer, sheet_name='Configuration', index=False)
                    self._adjust_column_widths(writer.sheets['Configuration'])
                except Exception as e:
                    logger.error(f"Error adding configuration data: {e}")

                try:
                    # Add cross-analytic detail data (all records from all analytics)
                    self._add_all_detail_data(writer, results_by_analytic)
                except Exception as e:
                    logger.error(f"Error adding detail data: {e}")

            logger.info(f"Generated consolidated department-level report: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Error generating consolidated main report: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None

    def _create_basic_summary(self, writer, results_by_analytic):
        """
        Create a basic summary when the detailed summary fails

        Args:
            writer: Excel writer
            results_by_analytic: Dictionary of results by analytic ID
        """
        # Create a simple summary with basic analytics information
        basic_data = [
            {'Analytics Summary': 'DEPARTMENT LEVEL QA ANALYTICS REPORT'},
            {'Analytics Summary': f"Date Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"},
            {'Analytics Summary': f"Number of Analytics: {len(results_by_analytic)}"},
            {'Analytics Summary': ""},
            {'Analytics Summary': "Analytics Processed:"}
        ]

        # Add each analytic
        for analytic_id, data in results_by_analytic.items():
            config = data['config']
            basic_data.append({'Analytics Summary': f"QA-{analytic_id}: {config['analytic_name']}"})

        # Create DataFrame and write to sheet
        basic_df = pd.DataFrame(basic_data)
        basic_df.to_excel(writer, sheet_name='Department Summary', index=False)

        # Add note about error
        try:
            worksheet = writer.sheets['Department Summary']
            row_count = len(basic_data) + 3  # Header row + data rows + buffer
            worksheet.cell(row=row_count,
                           column=1).value = "Note: Detailed summary could not be generated due to an error."
            worksheet.cell(row=row_count, column=1).font = writer.book.create_font(bold=True, color="FF0000")

            # Auto-adjust column widths
            self._adjust_column_widths(worksheet)
        except Exception as e:
            logger.error(f"Error formatting basic summary: {e}")
            # Continue without additional formatting

    def _create_department_summary(self, writer, results_by_analytic: Dict[str, Dict],
                                   leader_field_by_analytic: Dict[str, str],
                                   all_leaders: List[str]):
        """
        Create a department-level summary sheet

        Args:
            writer: Excel writer object
            results_by_analytic: Dictionary of results by analytic ID
            leader_field_by_analytic: Dictionary mapping analytic IDs to their leader field names
            all_leaders: List of all audit leaders
        """
        # Create header data
        header_data = [
            {'Overview': 'DEPARTMENT LEVEL QA ANALYTICS SUMMARY'},
            {'Overview': f"Date Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"},
            {'Overview': f"Number of Analytics: {len(results_by_analytic)}"},
            {'Overview': f"Number of Audit Leaders: {len(all_leaders)}"},
            {'Overview': ""},
            {'Overview': ""}
        ]

        # Create a pivot table style summary
        # First, gather data for each analytic and leader combination
        summary_data = []

        for analytic_id, data in results_by_analytic.items():
            config = data['config']
            results = data['results']
            analytic_name = config['analytic_name']
            threshold = config['thresholds']['error_percentage']

            if 'summary' in results and not results['summary'].empty:
                summary = results['summary']
                leader_field = leader_field_by_analytic[analytic_id]

                # For each leader in this analytic's summary
                for _, row in summary.iterrows():
                    leader = row[leader_field]
                    if pd.isna(leader):
                        continue

                    summary_row = {
                        'QA-ID': analytic_id,
                        'Analytic Name': analytic_name,
                        'Audit Leader': leader,
                        'GC': row.get('GC', 0),
                        'PC': row.get('PC', 0),
                        'DNC': row.get('DNC', 0),
                        'Total': row.get('Total', 0),
                        'DNC %': row.get('DNC_Percentage', 0),
                        'Threshold %': threshold,
                        'Exceeds Threshold': 'Yes' if row.get('DNC_Percentage', 0) > threshold else 'No'
                    }

                    summary_data.append(summary_row)

        # Create two summary tables:
        # 1. By Analytic (all leaders)
        # 2. By Leader (all analytics)

        if summary_data:
            # Convert to DataFrame
            summary_df = pd.DataFrame(summary_data)

            # Write header
            header_df = pd.DataFrame(header_data)
            header_df.to_excel(writer, sheet_name='Department Summary', index=False)
            worksheet = writer.sheets['Department Summary']

            # 1. Summary by Analytic
            analytic_pivot = pd.pivot_table(
                summary_df,
                values=['GC', 'PC', 'DNC', 'Total', 'DNC %'],
                index=['QA-ID', 'Analytic Name'],
                aggfunc={'GC': 'sum', 'PC': 'sum', 'DNC': 'sum', 'Total': 'sum', 'DNC %': 'mean'}
            ).reset_index()

            # Add thresholds
            analytic_pivot['Threshold %'] = analytic_pivot['QA-ID'].map(
                {id: data['config']['thresholds']['error_percentage'] for id, data in results_by_analytic.items()}
            )

            # Add exceeds threshold column
            analytic_pivot['Exceeds Threshold'] = analytic_pivot.apply(
                lambda x: 'Yes' if x['DNC %'] > x['Threshold %'] else 'No', axis=1
            )

            # Round percentages
            analytic_pivot['DNC %'] = analytic_pivot['DNC %'].round(2)

            # Write to sheet starting at an offset from header
            start_row = len(header_data) + 2
            worksheet.cell(row=start_row, column=1).value = "SUMMARY BY ANALYTIC"
            analytic_pivot.to_excel(
                writer,
                sheet_name='Department Summary',
                startrow=start_row + 1,
                index=False
            )

            # 2. Summary by Leader
            start_row = start_row + len(analytic_pivot) + 4
            worksheet.cell(row=start_row, column=1).value = "SUMMARY BY AUDIT LEADER"

            leader_pivot = pd.pivot_table(
                summary_df,
                values=['GC', 'PC', 'DNC', 'Total', 'DNC %'],
                index=['Audit Leader'],
                aggfunc={'GC': 'sum', 'PC': 'sum', 'DNC': 'sum', 'Total': 'sum', 'DNC %': 'mean'}
            ).reset_index()

            # Round percentages
            leader_pivot['DNC %'] = leader_pivot['DNC %'].round(2)

            # Calculate whether any analytic exceeds threshold for this leader
            leader_exceeds = summary_df.groupby('Audit Leader')['Exceeds Threshold'].apply(
                lambda x: 'Yes' if 'Yes' in x.values else 'No'
            ).reset_index()

            # Merge with leader_pivot
            leader_pivot = leader_pivot.merge(leader_exceeds, on='Audit Leader')

            leader_pivot.to_excel(
                writer,
                sheet_name='Department Summary',
                startrow=start_row + 1,
                index=False
            )

            # 3. Add a heatmap-style matrix (Leader x Analytic with DNC %)
            start_row = start_row + len(leader_pivot) + 4
            worksheet.cell(row=start_row, column=1).value = "DNC % HEATMAP BY LEADER AND ANALYTIC"

            # Create a flattened version of the heatmap data that avoids MultiIndex issues
            heatmap_data = []
            for _, row in summary_df.iterrows():
                heatmap_data.append({
                    'Audit Leader': row['Audit Leader'],
                    'QA-ID': row['QA-ID'],
                    'Analytic Name': row['Analytic Name'],
                    'DNC %': row['DNC %']
                })

            heatmap_df = pd.DataFrame(heatmap_data)

            # Create a pivot table that doesn't use MultiIndex columns
            try:
                # First approach: create a concatenated column for QA-ID and Analytic Name
                heatmap_df['Analytic'] = heatmap_df['QA-ID'] + ': ' + heatmap_df['Analytic Name']

                # Pivot with single-level column headers
                flat_heatmap = pd.pivot_table(
                    heatmap_df,
                    values='DNC %',
                    index=['Audit Leader'],
                    columns=['Analytic'],
                    aggfunc='mean'
                ).round(2)

                # Reset index to make Audit Leader a column
                flat_heatmap = flat_heatmap.reset_index()

                # Write to sheet
                flat_heatmap.to_excel(
                    writer,
                    sheet_name='Department Summary',
                    startrow=start_row + 1,
                    index=False
                )

                # Apply conditional formatting (color gradient) to the heatmap
                # Get reference to the worksheet
                worksheet = writer.sheets['Department Summary']

                # Determine the range of cells for the heatmap values (excluding headers and audit leader column)
                start_col = 2  # Column B (assuming Audit Leader is column A)
                start_data_row = start_row + 3  # First data row (+1 for header, +1 for startrow offset, +1 for the header row excel adds)
                end_row = start_data_row + len(flat_heatmap.index) - 1
                end_col = start_col + len(flat_heatmap.columns) - 1  # -1 because Audit Leader is already accounted for

                # Define color scale: green for low values, yellow for middle, red for high values
                from openpyxl.formatting.rule import ColorScaleRule
                from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

                # Apply color scale rule
                color_scale = ColorScaleRule(
                    start_type='num', start_value=0, start_color='63BE7B',  # Green
                    mid_type='num', mid_value=50, mid_color='FFEB84',  # Yellow
                    end_type='num', end_value=100, end_color='F8696B'  # Red
                )

                # Determine the cell range (e.g., "B4:C10")
                heatmap_range = f"{worksheet.cell(row=start_data_row, column=start_col).coordinate}:{worksheet.cell(row=end_row, column=end_col).coordinate}"
                worksheet.conditional_formatting.add(heatmap_range, color_scale)

                # Improve formatting: center all cells, add borders, make headers bold
                # Center all cells in the heatmap
                for row in range(start_data_row - 1, end_row + 1):  # Include header row
                    for col in range(1, end_col + 1):  # Include Audit Leader column
                        cell = worksheet.cell(row=row, column=col)
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                        # Add borders to all cells
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        cell.border = thin_border

                        # Make headers bold
                        if row == start_data_row - 1:  # Header row
                            cell.font = Font(bold=True)

                # Make the heatmap title bold and larger
                title_cell = worksheet.cell(row=start_row + 1, column=1)
                title_cell.font = Font(bold=True, size=12)

                # Add a note about the color coding
                note_row = end_row + 2
                note_cell = worksheet.cell(row=note_row, column=1)
                note_cell.value = "Note: Color coding indicates DNC % values - Green (0%) → Yellow (50%) → Red (100%)"
                note_cell.font = Font(italic=True)

                # Ensure the first column is wide enough for audit leader names
                worksheet.column_dimensions['A'].width = 25

                # Make analytic columns wider for better readability
                for col_idx in range(start_col, end_col + 1):
                    col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                    worksheet.column_dimensions[col_letter].width = 30

            except Exception as e:
                # Fallback approach if pivot table or formatting doesn't work
                logger.error(f"Error creating or formatting heatmap: {e}")

                # Create a simple table instead
                fallback_table = heatmap_df[['Audit Leader', 'Analytic', 'DNC %']]

                worksheet.cell(row=start_row + 1,
                               column=1).value = "Note: Simplified format due to Excel formatting constraints"

                fallback_table.to_excel(
                    writer,
                    sheet_name='Department Summary',
                    startrow=start_row + 3,
                    index=False
                )

            # Auto-adjust column widths
            self._adjust_column_widths(worksheet)
        else:
            # Create a basic summary if no data available
            basic_summary = pd.DataFrame(header_data)
            basic_summary.to_excel(writer, sheet_name='Department Summary', index=False)
            worksheet = writer.sheets['Department Summary']

            # Add message about no data
            start_row = len(header_data) + 2
            worksheet.cell(row=start_row, column=1).value = "No analytics data available for summary."

            self._adjust_column_widths(worksheet)

    def _add_all_detail_data(self, writer, results_by_analytic: Dict[str, Dict]):
        """
        Add all detail data to the report for cross-analytic analysis

        Args:
            writer: Excel writer object
            results_by_analytic: Dictionary of results by analytic ID
        """
        # Create a sheet with all detail records from all analytics
        all_detail_data = []

        for analytic_id, data in results_by_analytic.items():
            config = data['config']
            results = data['results']

            if 'detail' in results and not results['detail'].empty:
                # Add the analytic ID as a column for reference
                detail_df = results['detail'].copy()
                detail_df['QA_ID'] = analytic_id
                detail_df['Analytic_Name'] = config['analytic_name']

                # Add to our collection
                all_detail_data.append(detail_df)

        # If we have detail data, combine it and write to a sheet
        if all_detail_data:
            combined_df = pd.concat(all_detail_data, ignore_index=True)
            combined_df.to_excel(writer, sheet_name="All Detail Data", index=False)
            self._adjust_column_widths(writer.sheets["All Detail Data"])

        # Create an alternative heatmap visualization as a separate sheet for better readability
        try:
            self._create_enhanced_heatmap(writer, results_by_analytic)
        except Exception as e:
            logger.error(f"Error creating enhanced heatmap: {e}")

    def _create_enhanced_heatmap(self, writer, results_by_analytic):
        """
        Create an enhanced, more readable heatmap visualization

        Args:
            writer: Excel writer object
            results_by_analytic: Dictionary of results by analytic ID
        """
        # Collect data for all analytics and audit leaders
        heatmap_data = []

        for analytic_id, data in results_by_analytic.items():
            config = data['config']
            results = data['results']

            if 'summary' in results and not results['summary'].empty:
                summary = results['summary']
                leader_field = config['reporting']['group_by']

                for _, row in summary.iterrows():
                    leader = row[leader_field]
                    if pd.isna(leader):
                        continue

                    dnc_pct = row.get('DNC_Percentage', 0)
                    threshold = config['thresholds']['error_percentage']
                    exceeds = 'Yes' if dnc_pct > threshold else 'No'

                    heatmap_data.append({
                        'Audit Leader': leader,
                        'QA-ID': analytic_id,
                        'Analytic Name': config['analytic_name'],
                        'DNC %': dnc_pct,
                        'Threshold %': threshold,
                        'Exceeds Threshold': exceeds
                    })

        if not heatmap_data:
            return

        # Create DataFrame
        heatmap_df = pd.DataFrame(heatmap_data)

        # Create a new worksheet for the heatmap
        sheet_name = "DNC Heatmap"
        heatmap_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Get worksheet
        ws = writer.sheets[sheet_name]

        # Apply formatting
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.formatting.rule import ColorScaleRule

        # Define color fills for various thresholds
        green_fill = PatternFill(start_color="63BE7B", end_color="63BE7B", fill_type="solid")  # Green
        yellow_fill = PatternFill(start_color="FFEB84", end_color="FFEB84", fill_type="solid")  # Yellow
        orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Orange
        red_fill = PatternFill(start_color="F8696B", end_color="F8696B", fill_type="solid")  # Red

        # Define borders and alignment
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        bold_font = Font(bold=True)

        # Apply header formatting
        for cell in ws[1]:
            cell.font = bold_font
            cell.border = border
            cell.alignment = center_alignment

        # Format all data cells
        for row in ws.iter_rows(min_row=2, min_col=1, max_row=len(heatmap_data) + 1, max_col=len(heatmap_df.columns)):
            for cell in row:
                cell.border = border
                cell.alignment = center_alignment

                # Apply color to DNC % cells (column index 3 - zero-based, so 4 in Excel)
                if cell.column == 4:  # DNC % column
                    value = cell.value
                    if value is not None:
                        if value < 25:
                            cell.fill = green_fill
                        elif value < 50:
                            cell.fill = yellow_fill
                        elif value < 75:
                            cell.fill = orange_fill
                        else:
                            cell.fill = red_fill

                # Apply color to Exceeds Threshold cells
                if cell.column == 6:  # Exceeds Threshold column
                    if cell.value == 'Yes':
                        cell.fill = red_fill
                    else:
                        cell.fill = green_fill

        # Set column widths
        ws.column_dimensions['A'].width = 25  # Audit Leader
        ws.column_dimensions['B'].width = 10  # QA-ID
        ws.column_dimensions['C'].width = 35  # Analytic Name
        ws.column_dimensions['D'].width = 15  # DNC %
        ws.column_dimensions['E'].width = 15  # Threshold %
        ws.column_dimensions['F'].width = 20  # Exceeds Threshold

        # Add title and legend
        ws.insert_rows(1, 3)  # Add 3 rows at the top

        # Add title
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = "DNC PERCENTAGE HEATMAP BY AUDIT LEADER AND ANALYTIC"
        title_cell.font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add legend
        legend_cell = ws.cell(row=2, column=1)
        legend_cell.value = "Color Legend: "
        legend_cell.font = Font(bold=True)

        # Green legend
        green_cell = ws.cell(row=2, column=2)
        green_cell.value = "< 25%"
        green_cell.fill = green_fill
        green_cell.border = border
        green_cell.alignment = center_alignment

        # Yellow legend
        yellow_cell = ws.cell(row=2, column=3)
        yellow_cell.value = "25% - 49%"
        yellow_cell.fill = yellow_fill
        yellow_cell.border = border
        yellow_cell.alignment = center_alignment

        # Orange legend
        orange_cell = ws.cell(row=2, column=4)
        orange_cell.value = "50% - 74%"
        orange_cell.fill = orange_fill
        orange_cell.border = border
        orange_cell.alignment = center_alignment

        # Red legend
        red_cell = ws.cell(row=2, column=5)
        red_cell.value = "≥ 75%"
        red_cell.fill = red_fill
        red_cell.border = border
        red_cell.alignment = center_alignment

    def _generate_leader_report(self, leader: str, results_by_analytic: Dict[str, Dict],
                                leader_field_by_analytic: Dict[str, str]) -> str:
        """
        Generate a consolidated report for a specific audit leader

        Args:
            leader: Name of the audit leader
            results_by_analytic: Dictionary of results by analytic ID
            leader_field_by_analytic: Dictionary mapping analytic IDs to their leader field names

        Returns:
            Path to the generated report
        """
        try:
            # Create filename for this leader
            safe_leader_name = str(leader).replace('/', '_').replace('\\', '_').replace(' ', '_')
            timestamp = datetime.datetime.now().strftime("%Y%m%d")
            filename = f"Consolidated_{safe_leader_name}_{timestamp}.xlsx"
            output_path = os.path.join(self.output_dir, filename)

            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Create executive summary
                self._create_executive_summary(writer, leader, results_by_analytic, leader_field_by_analytic)

                # Add each analytic's data to separate sheets
                for analytic_id, data in results_by_analytic.items():
                    config = data['config']
                    results = data['results']
                    leader_field = leader_field_by_analytic[analytic_id]

                    # Filter data for this leader
                    if 'detail' in results and leader_field in results['detail'].columns:
                        # Filter detail data for this leader
                        leader_detail = results['detail'][results['detail'][leader_field] == leader]

                        if not leader_detail.empty:
                            # Write this analytic's detail to a sheet
                            sheet_name = f"QA-{analytic_id} Detail"
                            leader_detail.to_excel(writer, sheet_name=sheet_name, index=False)

                            # Auto-adjust column widths
                            self._adjust_column_widths(writer.sheets[sheet_name])

                    # Add configuration info for this analytic
                    config_sheet_name = f"QA-{analytic_id} Config"
                    config_data = self._create_config_data(data['config'], data['source_file'])
                    pd.DataFrame(config_data).to_excel(writer, sheet_name=config_sheet_name, index=False)
                    self._adjust_column_widths(writer.sheets[config_sheet_name])

            logger.info(f"Generated consolidated report for {leader}: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Error generating consolidated report for {leader}: {e}")
            return None

    def _create_executive_summary(self, writer, leader: str, results_by_analytic: Dict[str, Dict],
                                  leader_field_by_analytic: Dict[str, str]):
        """
        Create an executive summary sheet with overview of all analytics results

        Args:
            writer: Excel writer object
            leader: Audit leader name
            results_by_analytic: Dictionary of results by analytic ID
            leader_field_by_analytic: Dictionary mapping analytic IDs to their leader field names
        """
        # Create data for summary table
        summary_data = []

        for analytic_id, data in results_by_analytic.items():
            config = data['config']
            results = data['results']
            leader_field = leader_field_by_analytic[analytic_id]

            # Extract summary for this leader if available
            if 'summary' in results and leader_field in results['summary'].columns:
                leader_summary = results['summary'][results['summary'][leader_field] == leader]

                if not leader_summary.empty:
                    row = leader_summary.iloc[0].to_dict()

                    # Create summary row
                    summary_row = {
                        'Analytic ID': analytic_id,
                        'Analytic Name': config['analytic_name'],
                        'GC': row.get('GC', 0),
                        'PC': row.get('PC', 0),
                        'DNC': row.get('DNC', 0),
                        'Total': row.get('Total', 0),
                        'DNC %': row.get('DNC_Percentage', 0),
                        'Threshold %': config['thresholds']['error_percentage'],
                        'Exceeds Threshold': row.get('Exceeds_Threshold', False)
                    }

                    summary_data.append(summary_row)

        # Create DataFrame from summary data
        if summary_data:
            summary_df = pd.DataFrame(summary_data)

            # Add a header section with overall information
            header_data = [
                {'Overview': f"Consolidated QA Analytics Report"},
                {'Overview': f"Audit Leader: {leader}"},
                {'Overview': f"Date Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"},
                {'Overview': f"Number of Analytics: {len(results_by_analytic)}"},
                {'Overview': ""},  # Empty row as separator
                {'Overview': ""}  # Empty row as separator
            ]

            header_df = pd.DataFrame(header_data)

            # Write header and summary to executive summary sheet
            header_df.to_excel(writer, sheet_name='Executive Summary', index=False)

            # Get the worksheet to determine where to start the summary table
            worksheet = writer.sheets['Executive Summary']
            start_row = len(header_data) + 2  # +2 for header row and zero-indexing

            # Write summary data starting at the calculated row
            summary_df.to_excel(writer, sheet_name='Executive Summary',
                                startrow=start_row, index=False)

            # Auto-adjust column widths
            self._adjust_column_widths(worksheet)
        else:
            # Create a basic summary if no data available
            basic_summary = pd.DataFrame([
                {'Overview': f"Consolidated QA Analytics Report"},
                {'Overview': f"Audit Leader: {leader}"},
                {'Overview': f"Date Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"},
                {'Overview': f"No analytics data available for this audit leader."}
            ])

            basic_summary.to_excel(writer, sheet_name='Executive Summary', index=False)
            self._adjust_column_widths(writer.sheets['Executive Summary'])

    def _create_config_data(self, config: Dict, source_file: str = None) -> List[Dict]:
        """
        Create configuration data for display in Excel

        Args:
            config: Configuration dictionary
            source_file: Source data file path

        Returns:
            List of parameter/value pairs for display
        """
        config_data = []

        # Basic analytic information
        config_data.append({'Parameter': 'Analytic ID', 'Value': config['analytic_id']})
        config_data.append({'Parameter': 'Analytic Name', 'Value': config['analytic_name']})

        # Add description if available
        if 'analytic_description' in config:
            config_data.append({'Parameter': 'Description', 'Value': config['analytic_description']})

        # Add source file information
        if source_file:
            config_data.append({'Parameter': 'Source File', 'Value': os.path.basename(source_file)})

        # Threshold information
        config_data.append({'Parameter': 'Error Threshold (%)', 'Value': config['thresholds']['error_percentage']})
        if 'rationale' in config['thresholds']:
            config_data.append({'Parameter': 'Threshold Rationale', 'Value': config['thresholds']['rationale']})

        # Add validation rules section header
        config_data.append({'Parameter': '--- VALIDATION RULES ---', 'Value': ''})

        # Add validation rules information
        for i, validation in enumerate(config.get('validations', []), 1):
            rule_name = validation.get('rule', 'Unknown')
            description = validation.get('description', 'No description')

            config_data.append({'Parameter': f'Rule {i}', 'Value': f"{rule_name}: {description}"})

            # Add rationale if available
            if 'rationale' in validation:
                config_data.append({'Parameter': f'Rule {i} Rationale', 'Value': validation['rationale']})

        # Add metadata if available
        if 'report_metadata' in config:
            config_data.append({'Parameter': '--- REPORT METADATA ---', 'Value': ''})

            # Add all metadata fields
            for key, value in config['report_metadata'].items():
                # Convert key from snake_case to Title Case
                display_key = ' '.join(word.capitalize() for word in key.split('_'))
                config_data.append({'Parameter': display_key, 'Value': value})

        return config_data

    def _adjust_column_widths(self, worksheet):
        """
        Auto-adjust column widths in a worksheet

        Args:
            worksheet: Worksheet object to adjust
        """
        for idx, col in enumerate(worksheet.columns, 1):
            max_length = 0
            column = worksheet.cell(row=1, column=idx).column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = min(adjusted_width, 50)