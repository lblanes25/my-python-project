# Fixed version of report_generator.py
# Make sure to include the proper imports at the top

import os
import datetime
import pandas as pd
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from logging_config import setup_logging

logger = setup_logging()


class ReportGenerator:
    """Generates Excel reports from processed data"""

    def __init__(self, config: Dict, results: Dict):
        """
        Initialize report generator

        Args:
            config: Configuration dictionary
            results: Dictionary with 'detail' and 'summary' DataFrames
        """
        self.config = config
        self.results = results
        self.output_dir = "output"

        # Create output directory if it doesn't exist
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

    def generate_main_report(self, output_path: str = None, source_file: str = None) -> str:
        """
        Generate main report with all data

        Args:
            output_path: Optional path for output file
            source_file: Optional name of source data file

        Returns:
            Path to generated report
        """
        if not output_path:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"QA_{self.config['analytic_id']}_{timestamp}.xlsx"
            output_path = os.path.join(self.output_dir, filename)

        try:
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Write summary sheet
                self.results['summary'].to_excel(writer, sheet_name='Summary', index=False)

                # Write detail sheet
                self.results['detail'].to_excel(writer, sheet_name='Detail', index=False)

                # Create configuration data from config file
                config_data = self._create_config_sheet_data(source_file)

                # Write configuration data to sheet
                pd.DataFrame(config_data).to_excel(writer, sheet_name='Configuration', index=False)

                # Auto-adjust column widths
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    for idx, col in enumerate(worksheet.columns, 1):
                        max_length = 0
                        column = worksheet.cell(row=1, column=idx).column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[column].width = min(adjusted_width, 50)

            logger.info(f"Generated main report: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Error generating main report: {e}")
            return None

    def _create_config_sheet_data(self, source_file: str = None) -> List[Dict]:
        """
        Create detailed configuration data for the report

        Args:
            source_file: Optional source data file path

        Returns:
            List of dictionaries with Parameter/Value pairs
        """
        config_data = []

        # Basic analytic information
        config_data.append({'Parameter': 'Analytic ID', 'Value': self.config['analytic_id']})
        config_data.append({'Parameter': 'Analytic Name', 'Value': self.config['analytic_name']})

        # Add description if available
        if 'analytic_description' in self.config:
            config_data.append({'Parameter': 'Description', 'Value': self.config['analytic_description']})

        # Add run information
        config_data.append({'Parameter': 'Run Date', 'Value': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
        if source_file:
            config_data.append({'Parameter': 'Source File', 'Value': os.path.basename(source_file)})

        # Threshold information
        config_data.append({'Parameter': 'Error Threshold (%)', 'Value': self.config['thresholds']['error_percentage']})
        if 'rationale' in self.config['thresholds']:
            config_data.append({'Parameter': 'Threshold Rationale', 'Value': self.config['thresholds']['rationale']})

        # Add validation rules section header
        config_data.append({'Parameter': '--- VALIDATION RULES ---', 'Value': ''})

        # Add validation rules information
        for i, validation in enumerate(self.config.get('validations', []), 1):
            rule_name = validation.get('rule', 'Unknown')
            description = validation.get('description', 'No description')

            config_data.append({'Parameter': f'Rule {i}', 'Value': f"{rule_name}: {description}"})

            # Add rationale if available
            if 'rationale' in validation:
                config_data.append({'Parameter': f'Rule {i} Rationale', 'Value': validation['rationale']})

        # Add metadata section header if metadata exists
        if 'report_metadata' in self.config:
            config_data.append({'Parameter': '--- REPORT METADATA ---', 'Value': ''})

            # Add all metadata fields
            for key, value in self.config['report_metadata'].items():
                # Convert key from snake_case to Title Case
                display_key = ' '.join(word.capitalize() for word in key.split('_'))
                config_data.append({'Parameter': display_key, 'Value': value})

        # Add results section header
        config_data.append({'Parameter': '--- RESULTS SUMMARY ---', 'Value': ''})

        # Add results statistics
        if 'detail' in self.results and 'Compliance' in self.results['detail'].columns:
            detail = self.results['detail']
            total_records = len(detail)

            gc_count = sum(detail['Compliance'] == 'GC')
            dnc_count = sum(detail['Compliance'] == 'DNC')
            pc_count = sum(detail['Compliance'] == 'PC')

            config_data.append({'Parameter': 'Total Records', 'Value': total_records})
            config_data.append({'Parameter': 'Generally Conforms (GC)',
                                'Value': f"{gc_count} ({gc_count / total_records * 100:.1f}%)"})
            config_data.append({'Parameter': 'Does Not Conform (DNC)',
                                'Value': f"{dnc_count} ({dnc_count / total_records * 100:.1f}%)"})
            config_data.append({'Parameter': 'Partially Conforms (PC)',
                                'Value': f"{pc_count} ({pc_count / total_records * 100:.1f}%)"})

        return config_data

    def generate_individual_reports(self) -> List[str]:
        """
        Generate individual reports for each group, without showing data from other groups

        Returns:
            List of paths to generated reports
        """
        group_by_field = self.config['reporting']['group_by']
        report_paths = []

        if group_by_field not in self.results['detail'].columns:
            logger.error(f"Group field '{group_by_field}' not found in data")
            return report_paths

        # Get unique groups
        groups = self.results['detail'][group_by_field].unique()

        for group in groups:
            if pd.isna(group):
                continue  # Skip null values

            try:
                # Filter data for this group
                group_detail = self.results['detail'][self.results['detail'][group_by_field] == group]
                group_summary = self.results['summary'][self.results['summary'][group_by_field] == group]

                # Create filename
                safe_group_name = str(group).replace('/', '_').replace('\\', '_')
                timestamp = datetime.datetime.now().strftime("%Y%m%d")
                filename = f"QA_{self.config['analytic_id']}_{safe_group_name}_{timestamp}.xlsx"
                output_path = os.path.join(self.output_dir, filename)

                # Create Excel writer
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    # Write summary sheet
                    if not group_summary.empty:
                        group_summary.to_excel(writer, sheet_name='Summary', index=False)

                    # Write detail sheet
                    if not group_detail.empty:
                        group_detail.to_excel(writer, sheet_name='Detail', index=False)

                    # Create configuration data for this group
                    config_data = self._create_config_sheet_data()

                    # Add group-specific information
                    # Find where the section header is
                    for i, row in enumerate(config_data):
                        if row['Parameter'] == '--- RESULTS SUMMARY ---':
                            # Insert group information before the results
                            config_data.insert(i, {'Parameter': f'{group_by_field}', 'Value': group})
                            break

                    # Write configuration data
                    pd.DataFrame(config_data).to_excel(writer, sheet_name='Configuration', index=False)

                    # Auto-adjust column widths
                    for sheet_name in writer.sheets:
                        worksheet = writer.sheets[sheet_name]
                        for idx, col in enumerate(worksheet.columns, 1):
                            max_length = 0
                            column = worksheet.cell(row=1, column=idx).column_letter
                            for cell in col:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                    pass
                            adjusted_width = (max_length + 2)
                            worksheet.column_dimensions[column].width = min(adjusted_width, 50)

                report_paths.append(output_path)
                logger.info(f"Generated individual report for {group}: {output_path}")

            except Exception as e:
                logger.error(f"Error generating report for {group}: {e}")

        return report_paths